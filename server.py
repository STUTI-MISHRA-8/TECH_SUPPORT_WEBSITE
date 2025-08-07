from flask import Flask, request, jsonify, stream_with_context, Response
from flask_cors import CORS
import subprocess
import openpyxl
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

EXCEL_FILE = "unsolved.xlsx"

# Load unsolved issues from Excel file
def load_unsolved():
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        issues = []
        # Assuming first row is header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Ensure issue exists
                issues.append({"issue": row[0], "solution": row[1] or ""})
        return issues
    else:
        return []

# Save unsolved issues to Excel file
def save_unsolved(unsolved_issues):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Issue", "Solution"])  # Header row
    for issue in unsolved_issues:
        sheet.append([issue["issue"], issue["solution"]])
    workbook.save(EXCEL_FILE)

unsolved_issues = load_unsolved()

@app.route('/generate', methods=['POST'])
def generate():
    data = request.get_json()
    prompt = data['prompt']

    def generate_response():
        try:
            process = subprocess.Popen(['ollama', 'run', 'llama2', prompt], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            for line in process.stdout:
                yield f"data: {line}\n\n"
        except Exception as e:
            yield f"data: Error: {str(e)}\n\n"
        finally:
            if process:
                process.kill()

    return Response(stream_with_context(generate_response()), mimetype='text/event-stream')

@app.route('/forward', methods=['POST'])
def forward():
    data = request.get_json()
    issue = data['issue']

    new_issue = {"issue": issue, "solution": ""}
    unsolved_issues.append(new_issue)
    save_unsolved(unsolved_issues)

    return jsonify({"message": "Issue forwarded successfully"})

@app.route('/unsolved', methods=['GET'])
def get_unsolved():
    return jsonify(unsolved_issues)

@app.route('/solve', methods=['POST'])
def solve():
    data = request.get_json()
    issue_index = data['index']
    solution = data['solution']

    if 0 <= issue_index < len(unsolved_issues):
        unsolved_issues[issue_index]['solution'] = solution
        save_unsolved(unsolved_issues)
        return jsonify({"message": "Issue solved successfully"})
    else:
        return jsonify({"error": "Invalid issue index"}), 400

if __name__ == '__main__':
    app.run(debug=True, port=5000)